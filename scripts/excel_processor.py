import pandas as pd
from dataclasses import dataclass
from math import isnan
import openpyxl


@dataclass
class ExcelProcessor:
    def _select_form16(self, file_name: str, sheet_name: str = "FORM-16") -> None:
        """
        Opens an existing Excel workbook for Form-16 using openpyxl and selects the specified worksheet.

        This method loads the Excel file provided by `file_name` and sets the `self.form16` attribute
        to the loaded workbook object. It then selects the worksheet specified by `sheet_name` (defaulting
        to "FORM-16") and assigns it to the `self.ws` attribute for further processing.

        Args:
            file_name (str): The path to the Excel file to open. This should be a valid .xlsx file.
            sheet_name (str, optional): The name of the worksheet to select from the workbook.
            Defaults to "FORM-16".

        Raises:
            FileNotFoundError: If the specified Excel file does not exist.
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.form16 to the loaded openpyxl workbook.
            Sets self.ws to the selected worksheet within the workbook.
        """
        self.form16 = openpyxl.load_workbook(file_name)
        self._select_worksheet(self.form16, sheet_name)

    def _select_worksheet(self, workbook: openpyxl.Workbook, sheet_name: str) -> None:
        """
        Selects a worksheet from an openpyxl workbook by name and assigns it to self.ws.

        This method takes an openpyxl Workbook object and a worksheet name, and sets the
        self.ws attribute to the corresponding worksheet. It is used internally after loading
        a workbook to prepare for further processing or data manipulation.

        Args:
            workbook (openpyxl.Workbook): The loaded Excel workbook object.
            sheet_name (str): The name of the worksheet to select.

        Raises:
            KeyError: If the specified worksheet name does not exist in the workbook.

        Side Effects:
            Sets self.ws to the selected worksheet within the workbook.
        """
        self.ws = workbook[sheet_name]

    def _extract_details(self, file_path: str, sheet_name: str = "ITR Format") -> dict:
        """
        Extract various financial and personal details from multiple sheets of the provided Excel file.

        Reads the specified worksheet (default "ITR Format") and additional sheets such as
        "Home Loan", "Health Insurance", "Education Loan", and "Donation" to extract relevant information.
        Parses key-value pairs, lists, and specific cell values to build a dictionary containing all
        required details for Form-16 generation.

        Args:
            file_path (str): Path to the Excel file containing the data.
            sheet_name (str, optional): Worksheet name to extract general details from. Defaults to "ITR Format".

        Returns:
            dict: Extracted details, including

            #### General taxpayer info (from "ITR Format" sheet):
            - Keys: {field_name: value, ...}
            - List keys: {field_name: [value1, value2], ...}
            - Password: {"passwd": value}

            #### Home loan details (from "Home Loan" sheet):
            - Keys: "bank_name", "loan_ac_number", "date_of_sanction", "total_loan_amount",
              "loan_outstanding", "loan_interest", "bank_name2", "loan_ac_number2",
              "date_of_sanction2", "total_loan_amount2", "loan_outstanding2", "loan_interest2"

            #### Health insurance details (from "Health Insurance" sheet):
            - Keys: "HI_self_company_name", "HI_self_policy_number", "HI_self_premium_amount",
              "HI_self_company_name2", "HI_self_policy_number2", "HI_self_premium_amount2",
              "HI_parents_company_name", "HI_parents_policy_number", "HI_parents_premium_amount",
              "HI_parents_company_name2", "HI_parents_policy_number2", "HI_parents_premium_amount2"

            #### Education loan details (from "Education Loan" sheet):
            - Keys: "EL_bank_name", "EL_loan_ac_number", "EL_date_of_sanction", "EL_total_loan_amount",
              "EL_loan_outstanding", "EL_loan_interest", "EL_bank_name2", "EL_loan_ac_number2",
              "EL_date_of_sanction2", "EL_total_loan_amount2", "EL_loan_outstanding2", "EL_loan_interest2"

            #### Donation details (from "Donation" sheet):
            - Keys: "pan_of_donee", "name_of_donee", "address_of_donee", "donation_amount",
              "pan_of_donee2", "name_of_donee2", "address_of_donee2", "donation_amount2"

        Raises:
            FileNotFoundError: If the specified Excel file does not exist.
            ValueError: If any required sheet or cell is missing or cannot be read.

        Side Effects:
            Populates self.data with the extracted information.
        """
        self.data = {}

        print("\033[1;37m\033[1mStarting extraction...\033[0m\n")
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        # Extract key-value pairs from B5:C20 (Excel is 1-indexed, pandas is 0-indexed)
        for row in range(4, 20):  # B5 is row 4 (0-indexed), C20 is row 19
            key = df.iat[row, 1]  # Column B (0:Serial_No, so index 1)
            value = df.iat[row, 2]  # Column C (0:Serial_No, so index 2)

            # Only add to dictionary if both key and value are not NaN
            self.data[key] = value

        # Extract key-list pairs from B22:D51 (Excel is 1-indexed, pandas is 0-indexed)
        for row in range(21, 51):  # B22 is row 21, D51 is row 50
            key = df.iat[row, 1]  # Column B (index 1)
            val1 = df.iat[row, 2]  # Column C (index 2)
            val2 = df.iat[row, 3]  # Column D (index 3)
            self.data[key] = [val1, val2]
        # D13 is row 12 (0-indexed), column 3 (0-indexed)
        self.data["passwd"] = df.iat[12, 3]

        """ ################## Extracting Home Loan Details ################## """

        print("\033[1;33mExtracting Home Loan Details...\033[0m\n")

        # Extract Home Loan details from the "Home Loan" sheet
        # Assuming the Home Loan details are in a specific sheet named "Home Loan"
        df = pd.read_excel(file_path, sheet_name="Home Loan", header=None)

        # Home Loan Details 1st Bank
        self.data["HL_bank_name"] = df.iat[3, 1]
        self.data["HL_loan_ac_number"] = df.iat[3, 2]
        self.data["HL_date_of_sanction"] = df.iat[3, 3]
        self.data["HL_total_loan_amount"] = df.iat[3, 4]
        self.data["HL_loan_outstanding"] = df.iat[3, 5]
        self.data["HL_loan_interest"] = df.iat[3, 6]

        # Home Loan Details 2nd Bank
        self.data["HL_bank_name2"] = df.iat[4, 1]
        self.data["HL_loan_ac_number2"] = df.iat[4, 2]
        self.data["HL_date_of_sanction2"] = df.iat[4, 3]
        self.data["HL_total_loan_amount2"] = df.iat[4, 4]
        self.data["HL_loan_outstanding2"] = df.iat[4, 5]
        self.data["HL_loan_interest2"] = df.iat[4, 6]

        """ ################## Extracting Health Insurance Details ################## """

        print("\033[1;32mExtracting Health Insurance Details...\033[0m\n")

        # Extract Health Insurance details from the "Health Insurance" sheet
        # Assuming the Health Insurance details are in a specific sheet named "Health Insurance"
        df = pd.read_excel(file_path, sheet_name="Health Insurance", header=None)

        # Health Insurance Details for Self 1st Company
        self.data["HI_self_company_name"] = df.iat[2, 1]
        self.data["HI_self_policy_number"] = df.iat[2, 2]
        self.data["HI_self_premium_amount"] = df.iat[2, 3]

        # Health Insurance Details for Self 2nd Company
        self.data["HI_self_company_name2"] = df.iat[3, 1]
        self.data["HI_self_policy_number2"] = df.iat[3, 2]
        self.data["HI_self_premium_amount2"] = df.iat[3, 3]

        # Health Insurance Details for Parents
        self.data["HI_parents_company_name"] = df.iat[9, 1]
        self.data["HI_parents_policy_number"] = df.iat[9, 2]
        self.data["HI_parents_premium_amount"] = df.iat[9, 3]

        # Health Insurance Details for Parents 2nd Company
        self.data["HI_parents_company_name2"] = df.iat[10, 1]
        self.data["HI_parents_policy_number2"] = df.iat[10, 2]
        self.data["HI_parents_premium_amount2"] = df.iat[10, 3]

        """ ################## Extracting Education Loan Details ################## """

        print("\033[1;33mExtracting Education Loan Details...\033[0m\n")

        # Extract Education details from the "Education Loan" sheet
        # Assuming the Education Loan details are in a specific sheet named "Education Loan"
        df = pd.read_excel(file_path, sheet_name="Education Loan", header=None)

        # Education Loan Details 1st Bank
        self.data["EL_bank_name"] = df.iat[3, 1]
        self.data["EL_loan_ac_number"] = df.iat[3, 2]
        self.data["EL_date_of_sanction"] = df.iat[3, 3]
        self.data["EL_total_loan_amount"] = df.iat[3, 4]
        self.data["EL_loan_outstanding"] = df.iat[3, 5]
        self.data["EL_loan_interest"] = df.iat[3, 6]

        # Education Loan Details 2nd Bank
        self.data["EL_bank_name2"] = df.iat[4, 1]
        self.data["EL_loan_ac_number2"] = df.iat[4, 2]
        self.data["EL_date_of_sanction2"] = df.iat[4, 3]
        self.data["EL_total_loan_amount2"] = df.iat[4, 4]
        self.data["EL_loan_outstanding2"] = df.iat[4, 5]
        self.data["EL_loan_interest2"] = df.iat[4, 6]

        """ ################## Extracting Donation Details ################## """

        print("\033[1;33mExtracting Donation Details...\033[0m\n")

        # Extract Donation details from the "Donation" sheet
        # Assuming the Donation details are in a specific sheet named "Donation"
        df = pd.read_excel(file_path, sheet_name="Donation", header=None)

        # Donation Details 1st Organization
        self.data["pan_of_donee"] = df.iat[3, 1]
        self.data["name_of_donee"] = df.iat[3, 2]
        self.data["address_of_donee"] = df.iat[3, 3]
        self.data["donation_amount"] = df.iat[3, 4]

        # Donation Details 2nd Organization
        self.data["pan_of_donee2"] = df.iat[4, 1]
        self.data["name_of_donee2"] = df.iat[4, 2]
        self.data["address_of_donee2"] = df.iat[4, 3]
        self.data["donation_amount2"] = df.iat[4, 4]

        print("\033[1;32m\tDetails extracted successfully:\033[0m\n")

        for key, value in self.data.items():
            print(f"\033[1;34m\t{key}:\033[0m {value}")

        return self.data

    def create_form_16(self, itr_format: str, form_16: str) -> bool:
        """
        Create Form-16 from the given ITR format file.
        """
        try:

            """################ Detail Extraction ################"""
            # Extract details from the ITR format file
            details = self._extract_details(itr_format)

            """################ Form-16 Sheet ################"""
            # Load the Form-16 template
            self._select_form16(form_16, sheet_name="FORM-16")

            # Update the value of cell A1 in the worksheet
            # Write "Name, Designation, Department" in cell A1 using values from details dict
            name = details.get("Name", "")
            designation = details.get("Designation", "")
            department = details.get("Department/Company", "")
            self.ws["A1"] = f"{name}, {designation}, {department}".upper()

            """################ Income from Other Sources ################"""

            self.ws["C35"] = details.get("Interest on Saving A/c", "")[0]
            self.ws["C36"] = details.get("Interest on FD/RD/MIS", "")[0]

            """################ Deductions under 80C ################"""

            self.ws["F44"] = details.get("NPS PRAN No. (NPS Employee)", "")
            self.ws["F45"] = details.get("PF A/c No. (GPF/EPF Employee)", "")

            fields = [
                "LIC",
                "PPF",
                "SSY",
                "PLI",
                "Tuition Fees",
                "ELSS (Tax Saver Mutual Fund)",
                "ULIP",
                "NSC",
                "Senior Citizen Saving Scheme (SCSS)",
                "FD 05 Years (Tax Saving)",
                "Stamp Duty (Plot/Property)",
                "Home Loan Principal",
            ]
            for i, field in enumerate(fields):
                saving_name = f"C{49 + i}"
                document_number = f"F{49 + i}"

                self.ws[saving_name] = details.get(field, "")[0]
                self.ws[document_number] = details.get(field, "")[1]

            """################ 80CCD(1B) -NPS Employee Contribution ################"""

            # NPS PRAN No. (NPS Employee)
            self.ws["F63"] = details.get("NPS PRAN No. (NPS Employee)", "")

            """################ 80D - Deductions for Medical Expenses ################"""

            # Preventive Health Checkup Expenses for Employee and Family
            phc_self = details.get("Health Checkup Exp (Employee & family)", "")[0]

            if not isnan(phc_self):
                self.ws["C68"] = (
                    phc_self if ((phc_self <= 5000)) else 5000
                )
            else:
                self.ws["C68"] = 0

            # Preventive Health Checkup Expenses for Parents
            phc_parents = details.get(
                "Medical Exp (If Parents are Senior Citizen)", ""
            )[0]
            
            if not isnan(phc_parents):
                self.ws["C73"] = (
                    phc_parents if (phc_parents <= 50000) else 50000
                )
            else:
                self.ws["C73"] = 0

            """################ IT Calculation Sheet ################"""

            # Load the IT Calculation Sheet
            # Now the self.ws is already set to the "IT Calculation" sheet,
            self._select_worksheet(self.form16, sheet_name="IT Calculation")

            self.ws["D18"] = details.get("TDS/Tax Deducted", "")[0]

            """################ HRA Sheet ################"""

            # Load the HRA Sheet
            # Now the self.ws is already set to the "HRA" sheet,
            self._select_worksheet(self.form16, sheet_name="HRA")

            self.ws["C4"] = details.get("House Rent", "")[0]

            """################ Home Loan Sheet ################"""

            # Load the Home Loan Sheet
            # Now the self.ws is already set to the "Home Loan" sheet ( HL ),
            self._select_worksheet(self.form16, sheet_name="HL")

            # Home Loan Details 1st Bank
            self.ws["C4"] = details.get("HL_bank_name", "")
            self.ws["D4"] = details.get("HL_loan_ac_number", "")
            self.ws["E4"] = details.get("HL_date_of_sanction", "")
            self.ws["F4"] = details.get("HL_total_loan_amount", "")
            self.ws["G4"] = details.get("HL_loan_outstanding", "")
            self.ws["H4"] = details.get("HL_loan_interest", "")

            # Home Loan Details 2nd Bank
            self.ws["C5"] = details.get("HL_bank_name2", "")
            self.ws["D5"] = details.get("HL_loan_ac_number2", "")
            self.ws["E5"] = details.get("HL_date_of_sanction2", "")
            self.ws["F5"] = details.get("HL_total_loan_amount2", "")
            self.ws["G5"] = details.get("HL_loan_outstanding2", "")
            self.ws["H5"] = details.get("HL_loan_interest2", "")

            """################ Education Loan Sheet ################"""

            # Load the Education Loan Sheet
            # Now the self.ws is already set to the "Education Loan" sheet ( EL ),
            self._select_worksheet(self.form16, sheet_name="EL")

            # Education Loan Details 1st Bank
            self.ws["C4"] = details.get("EL_bank_name", "")
            self.ws["D4"] = details.get("EL_loan_ac_number", "")
            self.ws["E4"] = details.get("EL_date_of_sanction", "")
            self.ws["F4"] = details.get("EL_total_loan_amount", "")
            self.ws["G4"] = details.get("EL_loan_outstanding", "")
            self.ws["H4"] = details.get("EL_loan_interest", "")

            # Education Loan Details 2nd Bank
            self.ws["C5"] = details.get("EL_bank_name2", "")
            self.ws["D5"] = details.get("EL_loan_ac_number2", "")
            self.ws["E5"] = details.get("EL_date_of_sanction2", "")
            self.ws["F5"] = details.get("EL_total_loan_amount2", "")
            self.ws["G5"] = details.get("EL_loan_outstanding2", "")
            self.ws["H5"] = details.get("EL_loan_interest2", "")

            """################ Health Insurance Sheet ################"""

            # Load the Health Insurance Sheet
            # Now the self.ws is already set to the "Health Insurance" sheet ( HI ),
            self._select_worksheet(self.form16, sheet_name="HI")

            # Health Insurance Details for Self 1st Company
            self.ws["B4"] = details.get("HI_self_company_name", "")
            self.ws["C4"] = details.get("HI_self_policy_number", "")
            self.ws["D4"] = details.get("HI_self_premium_amount", "")

            # Health Insurance Details for Self 2nd Company
            self.ws["B5"] = details.get("HI_self_company_name2", "")
            self.ws["C5"] = details.get("HI_self_policy_number2", "")
            self.ws["D5"] = details.get("HI_self_premium_amount2", "")

            # Health Insurance Details for Self 1st Company
            self.ws["B11"] = details.get("HI_parents_company_name", "")
            self.ws["C11"] = details.get("HI_parents_policy_number", "")
            self.ws["D11"] = details.get("HI_parents_premium_amount", "")

            # Health Insurance Details for Self 2nd Company
            self.ws["B12"] = details.get("HI_parents_company_name2", "")
            self.ws["C12"] = details.get("HI_parents_policy_number2", "")
            self.ws["D12"] = details.get("HI_parents_premium_amount2", "")

            """################ Donation Sheet ################"""

            # Load the Donation Sheet
            # Now the self.ws is already set to the "Donation" sheet ( Donation ),
            self._select_worksheet(self.form16, sheet_name="Donation")

            # Donation Details for the 1st Donee
            self.ws["B3"] = details.get("name_of_donee", "")
            self.ws["C3"] = details.get("address_of_donee", "")
            self.ws["D3"] = details.get("pan_of_donee", "")
            self.ws["E3"] = details.get("donation_amount", "")

            # Donation Details for the 2nd Donee
            self.ws["B4"] = details.get("name_of_donee2", "")
            self.ws["C4"] = details.get("address_of_donee2", "")
            self.ws["D4"] = details.get("pan_of_donee2", "")
            self.ws["E4"] = details.get("donation_amount2", "")

            # Saving the Form-16 workbook
            self.form16.save(form_16)
            self.form16.close()

            return True
        except Exception as e:
            print(f"Error creating Form-16: {e}")
            return False


if __name__ == "__main__":
    # Create an instance of CSVProcessor
    test = ExcelProcessor()

    test.create_form_16(
        itr_format="form-16_generator/test/ITR Format (PIC).xlsx",
        form_16="form-16_generator/test/Form-16.xlsx",
    )
